import openpyxl

print("Processing Daily File.......")
#print("======================")
daily_mstr_file="F:/HEMA/STFC/27Jun2022/pdf24_merged_daily.xlsx"
mwb= openpyxl.load_workbook(daily_mstr_file)
no_of_sheets=len(mwb.sheetnames)
########################### IDENTIFY MULTIPLE EXECUTIVE SHEET ###########################
print("Executives - Summary")
print("--------------------")
sheets=mwb.sheetnames
exec_count=0
for i in sheets:
    sheet=mwb[i]
    #print("Sheet Name: ",sheet)
    arr=[]
    #print("Array declared")
    for c in sheet['B']:
        if c.value[:7]=="ZTotal:":
            arr.append(c.value)
    print(i,"--",len(arr))
    if len(arr)>1:multi_sheet=i
    exec_count=exec_count+len(arr)
print("Total Executives: ",exec_count)
print('Number of Sheets'
      ': ',no_of_sheets)
print("Multiple executives in sheet:",multi_sheet)
########################### END OF IDENTIFY MULTIPLE EXECUTIVE SHEET ###########################

for i in sheets:
    if(i!=multi_sheet):print(i,"hi")